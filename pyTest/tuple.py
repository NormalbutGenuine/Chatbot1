# numbers = (1, 2, 3, 4 ,5)
# new_numbers = numbers + (6,7)  # 튜플에 값을 추가하는 방법.
# print(new_numbers)
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '회원정보'
header_titles = ['단어', '품사']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

members = [('엔엘피', 'NNG'), ('나는 내일, 어제의 너와 만난다.', 'NNG')]
row_num = 2                                   # 엑셀 파일 만드는 법
for r, member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('user_dic.tsv')
wb.close()